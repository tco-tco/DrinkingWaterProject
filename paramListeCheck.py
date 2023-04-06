from analiste import ExcelFormat
from analiste import ExcelGelen

import openpyxl



path="C:/Users/username/Desktop/YÖNETMELİK PARAMETRELER.xlsx"
sheet="Sayfa1"
pathgelen="C:/Users/username/Desktop/param.xlsx"
sheetgelen="Sayfa1"

listeFormat=ExcelFormat(path,sheet)
listeFormatt=listeFormat.paramListesiFormat()

listeGelen=ExcelGelen(pathgelen,sheetgelen)
listeGelenn=listeGelen.paramListesiIl()

uygunListe=[]
uygunOlmayanListe=[]
elselist=[]


formatListe=[]
gelenListe=[]
for gelen in listeGelenn:
    gelenListe.append(str(gelen).lower())
    if str(gelen).endswith("ve bileşikleri"):
        gelenBilesikAyırma=str(gelen).split(" ")
        gelenListe.append(str(gelenBilesikAyırma[0]))
        print(str(gelenBilesikAyırma[0]))
    elif str(gelen).startswith("Toplam"):
        gelenBilesikAyırma=str(gelen).split(" ")
        gelenListe.append(str(gelenBilesikAyırma[1]))
        print(str(gelenBilesikAyırma[1]))

for format in listeFormatt:
    formatListe.append(str(format).lower())
                
sonliste=list(set(formatListe).intersection(gelenListe))
print(len(sonliste))
print(sonliste)
print("################################################")
for i in gelenListe:
    if i in formatListe:
        indexFormat=formatListe.index(i)
        indexGelen=gelenListe.index(i)
        workbook=openpyxl.load_workbook("C:\\Users\\username\\Desktop\\param.xlsx")

        sheet1=workbook["Sayfa1"]
        sheet1.cell(row=indexGelen+1,column=2, value=indexFormat)
        workbook.save("C:\\Users\\username\\Desktop\\param.xlsx") 



#########################################################

formatlaUyumsuzListe=[]
gelenleUyumsuzListe=[]

for format in formatListe:
    if format not in sonliste:
        formatlaUyumsuzListe.append(format)
        #print(format)
formatlaUyumsuzListe.sort()
print(formatlaUyumsuzListe)

for gelen in gelenListe:
    if gelen not in sonliste:
        gelenleUyumsuzListe.append(gelen)
        #print(gelen)

gelenleUyumsuzListe.sort()
print(gelenleUyumsuzListe)



karakterFormat=[]
karakterGelen=[]

karakter=[]
for i in formatlaUyumsuzListe:
    for char in '?\/:*"><|-,;':  
        if char in i:
            #print(f"{char} konuda var")
            karakter.append(char)


    konuYazim=[]
    for item in i:
        konuYazim.append(item)
    #print(konuYazim)

    for kar in karakter:
        for i in konuYazim:
            if kar==i:
                indexKar=konuYazim.index(kar)
                konuYazim.pop(indexKar)
                konuYazim.insert(indexKar,"")

    gelenHarfListe=[]
    for harf in konuYazim:
        if harf!=" ":
            gelenHarfListe.append(harf)
    #print(gelenHarfListe)
    formatJoin="".join(gelenHarfListe)
    print(formatJoin)

    karakterFormat.append(formatJoin)


#####

karakter=[]
for i in gelenleUyumsuzListe:
    for char in '?\/:*"><|-,;':  
        if char in i:
            #print(f"{char} konuda var")
            karakter.append(char)


    konuYazim=[]
    for item in i:
        konuYazim.append(item)
    #print(konuYazim)

    for kar in karakter:
        for i in konuYazim:
            if kar==i:
                indexKar=konuYazim.index(kar)
                konuYazim.pop(indexKar)
                konuYazim.insert(indexKar,"")

    gelenHarfListe=[]
    for harf in konuYazim:
        if harf!=" ":
            gelenHarfListe.append(harf)
    #print(gelenHarfListe)
    gelenJoin="".join(gelenHarfListe)
    print(gelenJoin)
    karakterGelen.append(gelenJoin)



sonliste1=list(set(karakterFormat).intersection(karakterGelen))
print(sonliste1)

for i in karakterGelen:
    if i in karakterFormat:
        indexFormat=formatListe.index(i)
        indexGelen=gelenListe.index(i)
        workbook=openpyxl.load_workbook("C:\\Users\\username\\Desktop\\param.xlsx")

        sheet1=workbook["Sayfa1"]
        sheet1.cell(row=indexGelen+1,column=2, value=indexFormat)
        workbook.save("C:\\Users\\username\\Desktop\\param.xlsx") 



##########################################






parantezUyumsuzFormat=[]
parantezUyumsuzGelen=[]

for i in formatlaUyumsuzListe:
    if str(i).endswith(")"):
        parant=str(i).split("(")
        parantezUyumsuzFormat.append(parant[0])
        print(parant[0])
for i in gelenleUyumsuzListe:
    if str(i).endswith(")"):
        parant=str(i).split("(")
        parantezUyumsuzGelen.append(parant[0])
        print(parant[0])


parantliste=list(set(parantezUyumsuzFormat).intersection(parantezUyumsuzGelen))
print(parantliste)
print(len(parantliste))

for i in parantezUyumsuzGelen:
    if i in parantezUyumsuzFormat:
        indexFormat=formatListe.index(i)
        indexGelen=gelenListe.index(i)
        workbook=openpyxl.load_workbook("C:\\Users\\username\\Desktop\\param.xlsx")

        sheet1=workbook["Sayfa1"]
        sheet1.cell(row=indexGelen+1,column=2, value=indexFormat)
        workbook.save("C:\\Users\\username\\Desktop\\param.xlsx") 

    


nihailiste=sonliste+sonliste1+parantliste
print(len(nihailiste))