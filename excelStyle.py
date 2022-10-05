import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from usbsUser import bilgisayarUsername


        # font = Font(name='Calibri',
        #          size=11,
        #          bold=False,
        #          italic=False,
        #          vertAlign=None,
        #          underline='none',
        #          strike=False,
        #          color='FF000000')

class ExcelDosya():
    def __init__(self,satir,sutun):
        self.satir=satir
        self.sutun=sutun



    def fillGreen(self,tesisAdi,kurum):
        workbook=openpyxl.load_workbook(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi}.xlsx")
        sheet1=workbook[f"{tesisAdi}"]
        sheet1.cell(row=self.satir,column=self.sutun).fill=PatternFill("solid",fgColor="0000FF00")
        thin = Side(border_style="thin", color="000000")
        sheet1.cell(row=self.satir,column=self.sutun).border = Border(top=thin, left=thin, right=thin, bottom=thin)
              
        workbook.save(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi}.xlsx")

    def fillYellow(self,tesisAdi,kurum):
        workbook=openpyxl.load_workbook(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi}.xlsx")
        sheet1=workbook[f"{tesisAdi}"]
        sheet1.cell(row=self.satir,column=self.sutun).fill=PatternFill("solid",fgColor="00FFFF00")
        thin = Side(border_style="thin", color="000000")
        sheet1.cell(row=self.satir,column=self.sutun).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        workbook.save(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi}.xlsx")

    def fillRed(self,tesisAdi,kurum):
        workbook=openpyxl.load_workbook(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi}.xlsx")
        sheet1=workbook[f"{tesisAdi}"]
        sheet1.cell(row=self.satir,column=self.sutun).fill=PatternFill("solid",fgColor="00FF0000")
        thin = Side(border_style="thin", color="000000")
        sheet1.cell(row=self.satir,column=self.sutun).border = Border(top=thin, left=thin, right=thin, bottom=thin)              

        workbook.save(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi}.xlsx")

    def fillGrey(self,tesisAdi,kurum):
        workbook=openpyxl.load_workbook(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi}.xlsx")
        sheet1=workbook[f"{tesisAdi}"]
        sheet1.cell(row=self.satir,column=self.sutun).fill=PatternFill("solid",fgColor="00C0C0C0")
        thin = Side(border_style="thin", color="000000")
        sheet1.cell(row=self.satir,column=self.sutun).border = Border(top=thin, left=thin, right=thin, bottom=thin)              

        workbook.save(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi}.xlsx")

