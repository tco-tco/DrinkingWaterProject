import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

class ExcelDosya():
    def __init__(self,satir,sutun,deger):
        self.workbook=openpyxl.load_workbook("C:/Users/username/Desktop/TESİSLERİM.xlsx")
        self.satir=satir
        self.sutun=sutun
        self.deger=deger

    def sıraYaz(self):           

        sheet1=self.workbook["Sayfa1"]
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger)
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).font=Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).border = Border(top=thin, left=thin, right=thin, bottom=thin)