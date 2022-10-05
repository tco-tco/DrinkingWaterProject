from selenium import webdriver
from selenium.webdriver.common.by import By

from selenium.webdriver.support.select import Select
from docx import Document

import openpyxl

import locale
locale.setlocale(locale.LC_ALL, 'tr_TR.utf8')
from usbsUser import username, password, bilgisayarUsername, ay,yil, ilListeOsman, ilListeSerhat, ilListesi

from excelStyle import ExcelDosya
import parametreFormat
import calcPandas

import os
import time
import os.path





class Usbs:
    def __init__(self):
        
        self.bilgisayarUsername=bilgisayarUsername

        self.browser=webdriver.Chrome(f"C:/Users/{bilgisayarUsername}/Desktop/python_temelleri_iş/ICMESUYU_PROJE/chromedriver.exe")
             
    
    def signIn(self):
        

        self.browser.get("https://usbs.tarimorman.gov.tr/verigirisi")
        self.browser.implicitly_wait(60)


        self.browser.find_element(By.XPATH,"//*[@id='UserName']").send_keys(username)
        self.browser.implicitly_wait(60)
        self.browser.find_element(By.XPATH,"//*[@id='Password']").send_keys(password)

        self.browser.implicitly_wait(60)
        time.sleep(1)
        self.browser.find_element(By.XPATH,"//*[@id='loginForm']/button").click()

        self.browser.implicitly_wait(60)
        time.sleep(1)
        
        sizeWindow=self.browser.get_window_size()
        print(sizeWindow)
        self.browser.implicitly_wait(60)
        self.browser.maximize_window() 
        self.browser.implicitly_wait(60)
        self.browser.find_element(By.XPATH,"/html/body/div[2]/div/button").click()
        self.browser.implicitly_wait(60)
        self.browser.find_element(By.XPATH,"/html/body/div[2]/div/ul/li[1]/a").click()
        self.browser.implicitly_wait(60)

        manuelOrAuto=input("""

        **********************************************************************************

        Manuel giriş için 'm', otomatik giriş için 'o' tuşuna ve 'Enter' tuşuna basınız... 

        **********************************************************************************
        """)
        print(manuelOrAuto)
        ilListe=[]
        if manuelOrAuto=="m":

            kacİlGirilecek=int(input("""
            ********************************************
                Kaç tane il girilecek?
            ************************************************
            """))            
            for i in range(1,kacİlGirilecek+1):
                ilAdiGiris=str(input(f"{i}. il adını giriniz: "))
                uppermap = {ord(u'ı'): u'I', ord(u'i'): u'İ'}
                ilAdiGiris = ilAdiGiris.translate(uppermap)
                ilAdiGiris=ilAdiGiris.upper()
                ilListe.append(ilAdiGiris)
            
            print("***************İl Listesi Tamam*******************")
        elif manuelOrAuto=="o":
            if ilListesi=="ilListeOsman":
                ilListe=ilListeOsman
            elif ilListesi=="ilListeSerhat":
                ilListe=ilListeSerhat

        print(ilListe)
        self.browser.switch_to.window(self.browser.window_handles[0])
        self.browser.implicitly_wait(60)
         
        for il in ilListe:
            print("############# YENİ İL #########################")
            print(il)
            
            uppermap = {ord(u'ı'): u'I', ord(u'i'): u'İ'}
            ilAdiGiris = il.translate(uppermap)
            il=ilAdiGiris.upper()

                        
            def selection(kurum):
                self.browser.switch_to.window(self.browser.window_handles[0])
                self.browser.implicitly_wait(60)
                selectil = Select(self.browser.find_element(By.XPATH,"//*[@id='KurumSec_KurumId']"))   
                self.browser.implicitly_wait(60)
                time.sleep(3)
                
                selectil.select_by_visible_text(kurum)
                self.browser.implicitly_wait(60)
                
            
            kurum=f"{str(il)} BÜYÜKŞEHİR BELEDİYESİ" 
            try:
                selection(kurum)
                
                hata=""
            
            except Exception as Ex:
                print(Ex)
                hata=f"Message: Could not locate element with visible text: {kurum}"
                print(hata)

            if hata==f"Message: Could not locate element with visible text: {kurum}":
                lowermap = {ord(u'I'): u'ı', ord(u'İ'): u'i'}
                il = str(il).translate(lowermap)
                il=il.capitalize()                
                kurum=f"{str(il)} Büyükşehir Belediyesi"
                    
               
                try:
                    selection(kurum)
                    1
                    hata=""
                except Exception as Ex:
                    print(Ex)
                    hata=f"Message: Could not locate element with visible text: {kurum}"
                    print(hata)           
                

                if hata==f"Message: Could not locate element with visible text: {kurum}":
            
                    print("""
                    #######################################################################
                    (a) İl adında hata varsa il adını yazıp 'Enter'ı tuşlayın...

                    (b) Eğer il adının doğru olduğundan eminseniz ilgili kurum adının tam adını 
                    öğrenip 1 rakamını ve ardından 'Enter'ı tuşladıktan sonra ilgili kurumun
                    tam adını yazın ve 'Enter'ı tuşlayın...                
                    ########################################################################
                    """)
                    yeniKurum=input("(a) ya da (b) seçeneklerine uygun tuşlamaları yapınız: ")
                    print(yeniKurum)

                    if yeniKurum=="1":
                        il=input("Kurum tam adını yazınız(büyük/küçük harfe dikkat ederek): ")
                        self.browser.switch_to.window(self.browser.window_handles[0])
                        self.browser.implicitly_wait(60)
                        selectil = Select(self.browser.find_element(By.XPATH,"//*[@id='KurumSec_KurumId']"))   
                        self.browser.implicitly_wait(60)
                        time.sleep(3)
                        selectil.select_by_visible_text(f"{str(il)}")
                        self.browser.implicitly_wait(60)
                    else:
                        il=yeniKurum
                        kurum=f"{str(il)} BÜYÜKŞEHİR BELEDİYESİ"

                        try:
                            selection(kurum)
                            
                            hata=""
                        except Exception as Ex:
                            print(Ex)
                            hata=f"Message: Could not locate element with visible text: {kurum}"
                            print(hata)

                        if hata==f"Message: Could not locate element with visible text: {kurum}":
                            
                            lowermap = {ord(u'I'): u'ı', ord(u'İ'): u'i'}
                            il = str(il).translate(lowermap)
                            il=il.capitalize()

                            selectil = Select(self.browser.find_element(By.XPATH,"//*[@id='KurumSec_KurumId']"))   
                            self.browser.implicitly_wait(60)
                            time.sleep(3)
                            selectil.select_by_visible_text(f"{str(il)} Büyükşehir Belediyesi")
                            self.browser.implicitly_wait(60)
                            print("Küçük Harfe Çevrilerek yeniden denendi")


            try:
                os.mkdir(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}")
                parametreFormat.createDB(il)
            except FileExistsError as FEE:
                print(FEE)
                print("""
                ********************************************************
                KLASÖR ZATEN MEVCUT, KLASÖRÜ BAŞKA YERE TAŞIYINIZ YA DA SİLİNİZ...
                KLASÖRÜ BAŞKA YERE KALDIRDIYSANIZ 'y' TUŞUNA ve SONRA ENTER'A BASINIZ...
                ********************************************************
                """)
                print("""
                ********************************************************
                                        YA DA
                MEVCUT KLASÖRÜN EKSİKLERİNİ TAMAMLAMAK İÇİN "1" TUŞUNA ve SONRA ENTER'A BASINIZ...
                ********************************************************
                """)
                hata=input("Seçiminize uygun girişi yapınız:")
                print(hata)
                

                if hata=="y":
                    os.mkdir(f'C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}')  
                    time.sleep(1)
                    parametreFormat.createDB(il)
                    print("OK-mkdir")
                elif hata=="1":
                    print("mevcut klasörle ve db ile devam edilecek...")

            
            time.sleep(1)
            self.browser.switch_to.window(self.browser.window_handles[0])
            self.browser.implicitly_wait(60)
            self.browser.find_element(By.XPATH,"/html/body/div[4]/div[1]/div[3]/div").click() # içme suyu arıtma tesisi modülüne tıklar
            self.browser.implicitly_wait(60)
            tesisListesi=self.browser.find_elements(By.XPATH,"/html/body/div/div[2]/div/div/div/div/div[1]/div/div[1]/div/div")
            self.browser.implicitly_wait(60)
            print(len(tesisListesi))

            for i in range(1,(len(tesisListesi))+1):
                print(i)
                tesisAdi1=self.browser.find_element(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/div[1]/div/div[1]/div/div[{i}]/div/a/h4").text
                self.browser.implicitly_wait(60)
                #print(tesisAdi1)
                

                tesisAdiListe=(str(tesisAdi1).split())
                #print(tesisAdiListe)
                cıkacakListe=["Sil","Güncelle","Ölçümler","Kimlik", "Kartı"]
                for cik in cıkacakListe:
                    tesisAdiListe.remove(cik)
                tesisAdi1="_".join(tesisAdiListe)
                #print(tesisAdi1)

                karakter=[]
                for char in '?\/:*"><|-,.()':  
                    if char in tesisAdi1:
                        #print(f"{char} konuda var")
                        karakter.append(char)


                konuYazim=[]
                for item in tesisAdi1:
                    konuYazim.append(item)
                #print(konuYazim)

                for kar in karakter:
                    for x in konuYazim:
                        if kar==x:
                            indexKar=konuYazim.index(kar)
                            konuYazim.pop(indexKar)
                            konuYazim.insert(indexKar," ")
                gelenHarfListe=[]
                for harf in konuYazim:
                    if harf!=" ":
                        gelenHarfListe.append(harf)
                #print(gelenHarfListe)
                tesisAdi1="".join(gelenHarfListe)
               
                print(tesisAdi1)
                
                try:
                    parametreFormat.createTable(tesisAdi1,il)
                    hata=""
            
                except Exception as Ex:
                    print(Ex)
                    hata=f"_mysql_connector.MySQLInterfaceError: Table '{tesisAdi1}' already exists"
                    
                    print(hata)    

                if hata==f"_mysql_connector.MySQLInterfaceError: Table '{tesisAdi1}' already exists":
                    time.sleep(1)
                    print ("hata verdi, yeni tesisle devam edilecek")
                    self.browser.switch_to.window(self.browser.window_handles[0])
                    self.browser.implicitly_wait(60)
                    time.sleep(5)
                    pass   
                else:
                                    
                    time.sleep(1)
                    parametreFormat.yonetmelikTabloFormat(tesisAdi1,il)
                    time.sleep(1)
                    self.browser.switch_to.window(self.browser.window_handles[0])
                    self.browser.implicitly_wait(60)
                   
                    self.browser.find_element(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/div[1]/div/div[1]/div/div[{i}]/div/a/h4/span[3]").click()    #ölçümler butonunu tıklar
                    self.browser.implicitly_wait(60)
                    self.browser.switch_to.window(self.browser.window_handles[1])
                    self.browser.implicitly_wait(60)
                    numuneSayısı=self.browser.find_elements(By.XPATH,f"//*[@id='IATIzlemeOlcumTable']/tbody/tr")                                                                     
                    self.browser.implicitly_wait(60)
                    time.sleep(2)
                    #print(numuneSayısı)
                    print(f"numune sayısı {len(numuneSayısı)}")
                    time.sleep(20)
                    #print(i)
                    
                    if int(len(numuneSayısı))==0:
                        print("Ölçüm yok-0 numune")
                        
                    elif int(len(numuneSayısı))!=0:
                        print("0 değil")

                        numuneTarihListesi=[]
                        for i in range(1,(len(numuneSayısı)+1)):
                            numuneTarihi=self.browser.find_element(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/div/div[2]/div/table/tbody/tr[{i}]/td[4]").text
                            self.browser.implicitly_wait(60)
                            #print(numuneTarihi)
                            tarihYil=str(numuneTarihi).split(".")
                            if int(tarihYil[2])==yil:
                                if int(tarihYil[1])>=ay:
                                    numuneTarihiListe=str(numuneTarihi).split(".")
                                    numuneTarihi="_".join(numuneTarihiListe) 
                                    numuneTarihListesi.append((i,numuneTarihi))
                            elif int(tarihYil[2])>yil:
                                numuneTarihiListe=str(numuneTarihi).split(".")
                                numuneTarihi="_".join(numuneTarihiListe) 
                                numuneTarihListesi.append((i,numuneTarihi))
                        


                        parametreFormat.addColumnName(tesisAdi1,"GİRİŞ",il)
                        for num,tarihnum in numuneTarihListesi:
                            parametreFormat.addColumnName(tesisAdi1,tarihnum,il)

                        parametreFormat.addColumnName(tesisAdi1,"ÇIKIŞ",il)
                        for num,tarihnum in numuneTarihListesi:
                            parametreFormat.addColumnName(tesisAdi1,f"Ç_{tarihnum}",il)

                         

                                
                        for num,tarih in numuneTarihListesi:

                            self.browser.find_element(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/div/div[2]/div/table/tbody/tr[{num}]/td[5]/p/button").click() #güncelle butonu ile bir numune tarihini tıklar
                            self.browser.implicitly_wait(60)
                            tesisAdi=self.browser.find_element(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/form/div/div[1]/div/h3").text
                            self.browser.implicitly_wait(60)
                            tesisAdiListe=(str(tesisAdi).split())
                            cıkacakListe=["Geri","Dön","Ölçüm", "Güncelle" ]
                            for i in cıkacakListe:
                                tesisAdiListe.remove(i)
                            tesisAdi=" ".join(tesisAdiListe)
                            self.browser.find_element(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/form/div/div[2]/div[1]/div[1]/input").text
                            self.browser.implicitly_wait(60)
                            parametreSayisi=self.browser.find_elements(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/form/div/div[2]/div[3]/div[1]/table/tbody/tr")
                            self.browser.implicitly_wait(60)
                            print("**********************************")
                            print(tesisAdi)
                            print(tarih)


                            
                            for i in range(1,int(len(parametreSayisi))+1):
                                parametreAdi=self.browser.find_element(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/form/div/div[2]/div[3]/div[1]/table/tbody/tr[{i}]/td[3]").text
                                self.browser.implicitly_wait(60)
                                birimGiris=self.browser.find_element(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/form/div/div[2]/div[3]/div[1]/table/tbody/tr[{i}]/td[4]").text
                                self.browser.implicitly_wait(60)
                                girisValue=self.browser.find_element(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/form/div/div[2]/div[3]/div[1]/table/tbody/tr[{i}]/td[5]/div/input").get_attribute("value")                                            
                                self.browser.implicitly_wait(60)
                                cıkısValue=self.browser.find_element(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/form/div/div[2]/div[3]/div[1]/table/tbody/tr[{i}]/td[6]/div/input").get_attribute("value")                                            
                                self.browser.implicitly_wait(60)
                                option1=self.browser.find_element(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/form/div/div[2]/div[3]/div[1]/table/tbody/tr[{i}]/td[5]/div/select/option[2]").get_attribute("selected")
                                self.browser.implicitly_wait(60)                           
                                option2=self.browser.find_element(By.XPATH,f"/html/body/div/div[2]/div/div/div/div/form/div/div[2]/div[3]/div[1]/table/tbody/tr[{i}]/td[6]/div/select/option[2]").get_attribute("selected")
                                self.browser.implicitly_wait(60)      

                                yokListe=[]
                                yokListe2=[]
                                listem=[(34, '1,1-Dikloroetan'), (35, '1,2-dikloroetan'), (36, '2,4-d_isooktil_ester'), (37, '4-kloroanilin'), (38, 'Akrilamid'), (39, 'Alaklor'), (40, 'Aldrin'), (5, 'Alüminyum'), 
                                (6, 'Amonyum'), (92, 'Anatoxin'), (7, 'Antimon'), (8, 'Arsenik'), (41, 'Asetaklor;_2-kloro-N-(etoksimetil)-N-(2-etil-6-metilfenil)asetamid'), (9, 'Bakır'), (10, 'Baryum'), (42, 'Benzen'), 
                                (43, 'Benzo(a)piren'), (11, 'Berilyum'), (12, 'Bor'), (13, 'Bromat'), (44, 'Bromoksinil'), (45, 'Bromür'), (2, 'Bulanıklık'), (15, 'Çinko'), (14, 'Cıva_ve_bileşikleri'), 
                                (96, 'Cryptosporidium_ookist'), (93, 'Cylindrospermopsin'), (82, 'DDT_(toplam)'), (16, 'Demir'), (46, 'Di(2-etilhekzil)fitalat_(DEHP)'), (47, 'Dieldrin'), (49, 'Diklobenil'), 
                                (48, 'Dikloroasetik_asit'), (50, 'Diklorometan'), (51, 'Diklorvos'), (52, 'Dikofol'), (53, 'Diuron'), (54, 'Etilentiyoüre_(ETU);_İmidazolidin-2-tiyon;_Etilentiyoüre_(ETU)'), 
                                (97, 'Fekal_Koliform'), (98, 'Fekal_Streptekok'), (17, 'Fenoller_(Fenol indeksi)_Para_nitroanilin_4_aminoantipirin'), (55, 'Fentiyon'), (18, 'Florür'), (56, 'Hekzakloro-benzen'), 
                                (57, 'Hekzakloro-siklohekzan'), (58, 'Heptaklor'), (59, 'Heptaklor_epoksit'), (3, 'İletkenlik_(arazi)'), (3, 'İletkenlik_(Lab)'), (19, 'Kadmiyum_ve_bileşikleri'), (60, 'Kaptan'), 
                                (61, 'Karbendazim'), (63, 'Klordan'), (67, 'Kloroasetik_asit'), (62, 'Klorotalonil'), (64, 'Klorpirifos'), (20, 'Klorür'), (21, 'Kobalt'), (22, 'Krom'), (23, 'Kurşun_ve_bileşikleri'), 
                                (65, 'Linuron'), (24, 'Mangan'), (66, 'Metolaklor'), (94, 'Microcystin-LR'), (68, 'Naftalin'), (25, 'Nikel_ve_bileşikleri'), (26, 'Nitrat'), (27, 'Nitrit'), (69, 'Oktabromodifenileter'), 
                                (28, 'Orto_Fosfat'), (70, 'Parakuat'), (71, 'Paration'), (72, 'Paration-metil'), (73, 'Pebulate'), (74, 'Pendimetalin'), (75, 'Pentakloro-benzen'), (76, 'Pentakloro-fenol'), (1, 'pH'), 
                                (78, "Poliklorlubifeniller_(PCB_ler)"), (4, 'Renk_(Pt-Co)'), (95, 'Saksitoksin'), (29, 'Selenyum'), (79, 'Sipermetrin'), (30, 'Siyanür'), (31, 'Sodyum'), (32, 'Sülfat'), (80, 'Terbutrin'), 
                                (81, 'Tetrakloroetilen'), (99, 'Toplam_Koliform'), (33, 'Toplam_Organik Karbon_(TOK)'), (83, 'Toplam_Pestisit'), (84, 'Tribenuron-metil'), (85, 'Trifluralin'), (86, 'Trihalometanlar'), 
                                (87, 'Trikloroasetik_asit'), (88, 'Trikloroetilen_(TRI)'), (89, 'Triklosan'), (90, 'Vanadyum'), (91, 'Vinil_Klorür')]

                                
                                if (len(girisValue)!=0): 
                                    #print(len(girisValue))
                                #and (parametreAdi in listem): 
                                    parametreAdiListe=str(parametreAdi).split(" ")
                                    parametreAdi="_".join(parametreAdiListe)
                                    parametreAdiListe=str(parametreAdi).split("'")
                                    parametreAdi="_".join(parametreAdiListe)             
                                    girisValue=str(round(float(girisValue.replace('.','').replace(',','.')), 11))
                                    for a,b in listem:

                                        if (parametreAdi==b) or (parametreAdi in listem):
                                            #print(a)
                                            # print(option)
                                            # print(type(option))
                                            if option1=="true":                     #Eğer LOQ seçeneği seçilmişse giriş değeri 2'ye bölünerek işlemlere devam edilir...
                                                #print(option)
                                                girisValue=float(girisValue)/2
                                                #print(girisValue)
                                                parametreFormat.updateValueParam(a,tesisAdi1,tarih,parametreAdi,girisValue,il)
                                                print(f"G{a},{parametreAdi},{birimGiris},{girisValue}")
                                                #print("parametre değeri 2'ye bölündü")
                                            else:
                                                parametreFormat.updateValueParam(a,tesisAdi1,tarih,parametreAdi,girisValue,il)
                                                print(f"G{a},{parametreAdi},{birimGiris},{girisValue}")
                                            break
                                        

                                        elif (parametreAdi!=b) or (parametreAdi not in listem):
                                            # print(b)
                                            # print(parametreAdi)
                                            yokListe.append((parametreAdi,girisValue))
                                        
                                            
                                if (len(cıkısValue)!=0): 
                                    #print(len(girisValue))
                                #and (parametreAdi in listem): 
                                    parametreAdiListe=str(parametreAdi).split(" ")
                                    parametreAdi="_".join(parametreAdiListe)
                                    parametreAdiListe=str(parametreAdi).split("'")
                                    parametreAdi="_".join(parametreAdiListe)             
                                    cıkısValue=str(round(float(cıkısValue.replace('.','').replace(',','.')), 11))
                                    for a,b in listem:

                                        if (parametreAdi==b) or (parametreAdi in listem):

                                            if option2=="true":
                                                #print(option)
                                                cıkısValue=float(cıkısValue)/2
                                                #print(cıkısValue)
                                                parametreFormat.updateValueParam(a,tesisAdi1,f"Ç_{tarih}",parametreAdi,cıkısValue,il)
                                                print(f"Ç{a},{parametreAdi},{birimGiris},{cıkısValue}")
                                                #print("parametre değeri 2'ye bölündü")
                                            else:
                                                parametreFormat.updateValueParam(a,tesisAdi1,f"Ç_{tarih}",parametreAdi,cıkısValue,il)
                                                print(f"Ç{a},{parametreAdi},{birimGiris},{cıkısValue}")
                                            break

                                        elif (parametreAdi!=b) or (parametreAdi not in listem):
                                            # print(b)
                                            # print(parametreAdi)
                                            yokListe2.append((parametreAdi,cıkısValue))
                                        


                            #print(yokListe)  
                            setYokListe=set(yokListe)  
                            setYokListe2=set(yokListe2)  
                            print(setYokListe)    
                            if len(setYokListe)!=0:
                                file=open(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi1}.txt","a", encoding="utf-8") #tesis adında bir .txt dosyası oluşturur
                                file.write(f"{tarih}-Giriş-{setYokListe}\n")
                                file.write(f"{tarih}-Çıkış-{setYokListe2}\n")
                                file.close()
                            else:
                                print("yok listesi boş")  
                                    
                            print("1")
                            self.browser.back()
                            self.browser.implicitly_wait(60)
        
                        print("2")
                        self.browser.back()
                        self.browser.implicitly_wait(60)
                        
                    print("3")
                    
                    self.browser.close()
                    self.browser.implicitly_wait(60)
                    time.sleep(2)
                    self.browser.switch_to.window(self.browser.window_handles[0])  
                    self.browser.implicitly_wait(60)         
                    
                    print("OK2")  

                    
                    for id in range(1,100):
                        result=calcPandas.avarageCalc(tesisAdi1,id,il,(len(numuneSayısı)+1))
                        result2=calcPandas.avarageCalcCıkıs(tesisAdi1,id,il,(len(numuneSayısı)+1))
                        print(f"giriş {result}")
                        print(f"çıkış {result2}")
                        #print(type(result)) #numpy.float64
                        result = result.tolist()
                        #print(type(result))
                        result2 = result2.tolist()

                        parametreFormat.updateValue(id,tesisAdi1,"Giriş_Ortalama",result,il)   
                        parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Ortalama",result2,il)   
                        A1F=parametreFormat.readA1Value(id,tesisAdi1,il,1)
                        A2F=parametreFormat.readA1Value(id,tesisAdi1,il,2)
                        A3F=parametreFormat.readA1Value(id,tesisAdi1,il,3)
                        A1=float(A1F)
                        # print(type(A1)) #str  NoneType-->boşken                    
                        # print(A2F)
                        # print(A3F)
                        if id==1:
                            pass                    
                        elif (A2F!="-") and (A3F!="-"):
                            A2=float(A2F)
                            A3=float(A3F)

                            if result<=A1:
                                parametreFormat.updateValue(id,tesisAdi1,"Giriş_Sonuç","A1",il) 
                            elif (A1<result<=A2):
                                parametreFormat.updateValue(id,tesisAdi1,"Giriş_Sonuç","A2",il) 
                            elif (A2<result<=A3):
                                parametreFormat.updateValue(id,tesisAdi1,"Giriş_Sonuç","A3",il) 
                            elif (result>A3):
                                parametreFormat.updateValue(id,tesisAdi1,"Giriş_Sonuç","A3",il) 

                        elif (A2F=="-") and (A3F=="-"):                        
                            if result<=A1:
                                parametreFormat.updateValue(id,tesisAdi1,"Giriş_Sonuç","A1",il) 
                            elif result>A1:
                                parametreFormat.updateValue(id,tesisAdi1,"Giriş_Sonuç","A3",il)  

                        elif (A2F=="-"):    
                            A3=float(A3F)                    
                            if result<=A1:
                                parametreFormat.updateValue(id,tesisAdi1,"Giriş_Sonuç","A1",il) 
                            elif (A1<result<=A3):
                                parametreFormat.updateValue(id,tesisAdi1,"Giriş_Sonuç","A3",il) 
                            elif (result>A3):
                                parametreFormat.updateValue(id,tesisAdi1,"Giriş_Sonuç","A3",il) 
                                    
                        elif (A3F=="-"):         
                            A2=float(A2F)              
                            if result<=A1:
                                parametreFormat.updateValue(id,tesisAdi1,"Giriş_Sonuç","A1",il) 
                            elif (A1<result<=A2):
                                parametreFormat.updateValue(id,tesisAdi1,"Giriş_Sonuç","A2",il) 
                            elif (A2<result):
                                parametreFormat.updateValue(id,tesisAdi1,"Giriş_Sonuç","A3",il)                        

                        
                        time.sleep(30)
                        if id==1:
                            pass                    
                        elif (A2F!="-") and (A3F!="-"):
                            A2=float(A2F)
                            A3=float(A3F)

                            if result2<=A1:
                                parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Sonuç","A1",il) 
                            elif (A1<result2<=A2):
                                parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Sonuç","A2",il) 
                            elif (A2<result2<=A3):
                                parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Sonuç","A3",il) 
                            elif (result2>A3):
                                parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Sonuç","A3",il) 

                        elif (A2F=="-") and (A3F=="-"):                        
                            if result2<=A1:
                                parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Sonuç","A1",il) 
                            elif result2>A1:
                                parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Sonuç","A3",il)  

                        elif (A2F=="-"):    
                            A3=float(A3F)                    
                            if result2<=A1:
                                parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Sonuç","A1",il) 
                            elif (A1<result2<=A3):
                                parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Sonuç","A3",il) 
                            elif (result2>A3):
                                parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Sonuç","A3",il) 
                                    
                        elif (A3F=="-"):         
                            A2=float(A2F)              
                            if result2<=A1:
                                parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Sonuç","A1",il) 
                            elif (A1<result2<=A2):
                                parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Sonuç","A2",il) 
                            elif (A2<result2):
                                parametreFormat.updateValue(id,tesisAdi1,"Çıkış_Sonuç","A3",il)                        
                    print("giriş ortalama değerleri girildi")
                    print("çıkış ortalama değerleri girildi")

                    excelExists=os.path.exists(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi1}.xlsx")
                    if excelExists==True:
                        print("""
                        **************************************************************
                        LÜTFEN EXCEL DOSYASINI SİLİP, "y" TUŞUNA VE SONRA ENTER'A BASINIZ...
                        **************************************************************

                        """)
                        excelSilindi=input()
                        if excelSilindi=="y":
                            parametreFormat.createExcel(kurum,tesisAdi1)

                    parametreFormat.createExcel(kurum,tesisAdi1)
                    parametreFormat.WriteToExcel(il,kurum,tesisAdi1)              
                    print("excel tamam")
                    #sonuçlardaki renklendirme
                    workbook=openpyxl.load_workbook(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi1}.xlsx")
                    sheet1=workbook[f"{tesisAdi1}"]
                    for row in range(1,100):
                        girisSonuc=sheet1.cell(row=row+1, column=11).value
                        print(girisSonuc)
                        #print(type(girisSonuc))
                        if girisSonuc=="A1":
                            excel=ExcelDosya(int(row+1),11)
                            excel.fillGreen(tesisAdi1,kurum)
                        elif girisSonuc=="A2":
                            excel=ExcelDosya(int(row+1),11)
                            excel.fillYellow(tesisAdi1,kurum)
                        elif girisSonuc=="A3":
                            excel=ExcelDosya(int(row+1),11)
                            excel.fillRed(tesisAdi1,kurum)
                        elif girisSonuc=="-":
                            print("PASS")
                            pass
                        else:
                            print("PASS")
                            pass
                    for row in range(1,100):
                        cıkısSonuc=sheet1.cell(row=row+1, column=13).value
                        print(cıkısSonuc)
                        #print(type(cıkısSonuc))
                        if cıkısSonuc=="A1":
                            excel=ExcelDosya(int(row+1),13)
                            excel.fillGreen(tesisAdi1,kurum)
                        elif cıkısSonuc=="A2":
                            excel=ExcelDosya(int(row+1),13)
                            excel.fillYellow(tesisAdi1,kurum)
                        elif cıkısSonuc=="A3":
                            excel=ExcelDosya(int(row+1),13)
                            excel.fillRed(tesisAdi1,kurum)
                        elif cıkısSonuc=="-":
                            print("PASS -")
                            pass
                        else:
                            print("PASS")
                            pass
                    #değerlerdeki renklendirme
                    workbook=openpyxl.load_workbook(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi1}.xlsx")
                    sheet1=workbook[f"{tesisAdi1}"]
                    columns=sheet1.max_column
                    print(columns)

                
                    for row in range(1,100):
                        A1F=sheet1.cell(row=row+1, column=7).value
                        #print(type(A1F))    #str
                        A2F=sheet1.cell(row=row+1, column=8).value
                        print(A2F)
                        A3F=sheet1.cell(row=row+1, column=9).value
                        print(A3F)
                        for columnn in range(14,columns+1):
                            result=sheet1.cell(row=row+1, column=columnn).value
                            print(result)   #none

                            A1=float(A1F)
                            if row==1:
                                pass
                            elif result==None:
                                print("nonetype-pass")
                                pass

                            elif (A2F!="-") and (A3F!="-"):
                                A2=float(A2F)
                                A3=float(A3F)

                                if float(result)<=A1:
                                    pass
                                elif (A1<float(result)<=A2):
                                    excel=ExcelDosya(int(row+1),columnn)
                                    excel.fillYellow(tesisAdi1,kurum)
                                elif (A2<float(result)<=A3):
                                    excel=ExcelDosya(int(row+1),columnn)
                                    excel.fillRed(tesisAdi1,kurum) 
                                elif (float(result)>A3):
                                    excel=ExcelDosya(int(row+1),columnn)
                                    excel.fillRed(tesisAdi1,kurum)
                            elif (A2F=="-") and (A3F=="-"):                        
                                if float(result)<=A1:
                                    pass
                                elif float(result)>A1:
                                    excel=ExcelDosya(int(row+1),columnn)
                                    excel.fillRed(tesisAdi1,kurum) 

                            elif (A2F=="-"):    
                                A3=float(A3F)                    
                                if float(result)<=A1:
                                    pass
                                elif (A1<float(result)<=A3):
                                    excel=ExcelDosya(int(row+1),columnn)
                                    excel.fillRed(tesisAdi1,kurum)
                                elif (float(result)>A3):
                                    excel=ExcelDosya(int(row+1),columnn)
                                    excel.fillRed(tesisAdi1,kurum)
                                        
                            elif (A3F=="-"):         
                                A2=float(A2F)              
                                if float(result)<=A1:
                                    pass
                                elif (A1<float(result)<=A2):
                                    excel=ExcelDosya(int(row+1),columnn)
                                    excel.fillYellow(tesisAdi1,kurum)
                                elif (A2<float(result)):
                                    excel=ExcelDosya(int(row+1),columnn)
                                    excel.fillRed(tesisAdi1,kurum)             

                    workbook=openpyxl.load_workbook(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi1}.xlsx")
                    sheet1=workbook[f"{tesisAdi1}"]
                    sheet1.delete_cols(1, 2)
                    
                    print("OK-excel")
                    time.sleep(2)

                    sheet1.delete_cols(2, 1)
                    workbook.save(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi1}.xlsx")
                    print("OK-excel2")


            self.browser.back()
            self.browser.implicitly_wait(60) 
            self.browser.back()
            self.browser.implicitly_wait(60)       
            print("OK3")            
                                


        


usbs = Usbs()
usbs.signIn()