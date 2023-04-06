

from asyncio.windows_events import NULL
import mysql.connector
import openpyxl
from openpyxl import Workbook
import pandas as pd
import time
import pandas.io.sql as sql
from usbsUser import bilgisayarUsername

def createDB(ilAdi):
    connection=mysql.connector.connect(host="localhost", user= "root",password="password")
    cursor=connection.cursor()
    cursor.execute(f"CREATE DATABASE {ilAdi}")
def createTable(tesisAdi,ilAdi):
    connection=mysql.connector.connect(host="localhost", user= "root",password="password", database=f"{ilAdi}")
    cursor=connection.cursor()

    cursor.execute(f"CREATE TABLE {tesisAdi} (ID VARCHAR(3) NOT NULL, Parametre_Yonetmelik VARCHAR(50), Parametre_Ölçüm VARCHAR(200), CAS_NO VARCHAR(20), Birim VARCHAR(25), A1 VARCHAR(25), A2 VARCHAR(25), A3 VARCHAR(25), Giriş_Ortalama VARCHAR(50), Giriş_Sonuç VARCHAR(5), Çıkış_Ortalama VARCHAR(50), Çıkış_Sonuç VARCHAR(5))") #kayıt eklemek için kullanılır



def insertValues(tesisAdi,list,ilAdi):
    connection=mysql.connector.connect(host="localhost", user= "root",password="password", database=f"{ilAdi}")
    cursor=connection.cursor()

    sql=f"INSERT INTO {tesisAdi}(ID, Parametre_Yonetmelik, Parametre_Ölçüm, CAS_NO, Birim, A1, A2, A3, Giriş_Ortalama, Giriş_Sonuç, Çıkış_Ortalama, Çıkış_Sonuç) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"   # Products: tablo adı, parantez içine kolonların adını giriyoruz, VALUES parantez içinde ise değerler kolon sırasına uygun olarak girilir(%s'ler yer tutucu görevinde)
                                                                                       # Nullable eğer NO ise mutlaka değer girilmelii değilse tüm kolonların adını vermeye gerek olmadan işlem yapılabilir
    values=list

    cursor.executemany(sql,values)  # birden fazla kayıt ekleneceği zaman execute yerine executemany kullanılır

    try:
        connection.commit()
        print(f"{cursor.rowcount} tane kayıt eklendi")
        print(f"son eklenen kayıt id: {cursor.lastrowid}")
    except mysql.connector.Error as err:
        print("hata: ", err)



def insertValue(tesisAdi,ID, Parametre_Yonetmelik, Parametre_Ölçüm, CAS_NO, Birim, A1, A2, A3, Giriş_Ortalama, Giriş_Sonuç, Çıkış_Ortalama, Çıkış_Sonuç,ilAdi):
    connection=mysql.connector.connect(host="localhost", user= "root",password="password", database=f"{ilAdi}")
    cursor=connection.cursor()

    sql=f"INSERT INTO {tesisAdi}(ID, Parametre_Yonetmelik, Parametre_Ölçüm, CAS_NO, Birim, A1, A2, A3, Giriş_Ortalama, Giriş_Sonuç, Çıkış_Ortalama, Çıkış_Sonuç) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"   
                                                                                       # Nullable eğer NO ise mutlaka değer girilmelii değilse tüm kolonların adını vermeye gerek olmadan işlem yapılabilir
    values=(ID, Parametre_Yonetmelik, Parametre_Ölçüm, CAS_NO, Birim, A1, A2, A3, Giriş_Ortalama, Giriş_Sonuç, Çıkış_Ortalama, Çıkış_Sonuç)

    cursor.execute(sql,values)

    try:
        connection.commit()
        print(f"{cursor.rowcount} tane kayıt eklendi")

    except mysql.connector.Error as err:
        print("hata: ", err)


def addColumnName(tesisAdi,olcumTarih,ilAdi):
    connection=mysql.connector.connect(host="localhost", user= "root",password="password", database=f"{ilAdi}")
    cursor=connection.cursor()

    query =f"ALTER TABLE {tesisAdi} ADD {olcumTarih} VARCHAR(25)"
    
    cursor.execute(query)




def yonetmelikTabloFormat(tesisAdi,ilAdi):

    path=f"C:/Users/{bilgisayarUsername}/Desktop/YÖNETMELİK PARAMETRELER.xlsx"
    sheet="Sayfa1"

    sheet1=openpyxl.load_workbook(path)[sheet]



    
    for x in range(2,101):    
        ID=sheet1[f"A{x}"].value
        Parametre_Yonetmelik=sheet1[f"B{x}"].value
        CAS_NO=sheet1[f"C{x}"].value
        Birim=sheet1[f"D{x}"].value
        A1=sheet1[f"E{x}"].value
        A2=sheet1[f"F{x}"].value
        A3=sheet1[f"G{x}"].value   
        Parametre_Ölçüm="-"
        Giriş_Ortalama= "-"
        Giriş_Sonuç="-"
        Çıkış_Ortalama="-"
        Çıkış_Sonuç="-"

        insertValue(tesisAdi,ID, Parametre_Yonetmelik, Parametre_Ölçüm, CAS_NO, Birim, A1, A2, A3, Giriş_Ortalama, Giriş_Sonuç, Çıkış_Ortalama, Çıkış_Sonuç,ilAdi)




def updateValueParam(id,tesisAdi,tarih,parametre,value,ilAdi):

    connection=mysql.connector.connect(host="localhost", user= "root",password="password", database=f"{ilAdi}")
    cursor=connection.cursor()
    
    
    sql="Update {ad} Set Parametre_Ölçüm='{param}', {date}='{deger}' where ID='{no}'".format(ad=tesisAdi,param=parametre,date=tarih,deger=value,no=id)

    
    cursor.execute(sql)   

    try:
        connection.commit()

        
    except mysql.connector.Error as err:
        print("hata: ", err)

    finally:
        connection.close()
        #print("database bağlantısı kapandı")
def updateValue(id,tesisAdi,tarih,value,ilAdi):

    connection=mysql.connector.connect(host="localhost", user= "root",password="password", database=f"{ilAdi}")
    cursor=connection.cursor()
    
    
    sql="Update {ad} Set {date}='{deger}' where ID='{no}'".format(ad=tesisAdi,date=tarih,deger=value,no=id)
    #values=(parametre,value,id)
    #sql = "TRUNCATE TABLE `" + placeholder_variable + "`"
    
    cursor.execute(sql)   

    try:
        connection.commit()
        #print(f"{cursor.rowcount} tane kayıt güncellendi")
        
    except mysql.connector.Error as err:
        print("hata: ", err)

    finally:
        connection.close()
        #print("database bağlantısı kapandı")


def readA1Value(id,tesisAdi,ilAdi,x):
    connection=mysql.connector.connect(host="localhost", user= "root",password="password", database=f"{ilAdi}")
    cursor=connection.cursor()
    Ax="Select A{num} From {tesis} Where ID='{ID}'".format(tesis=tesisAdi,ID=id,num=x)

    cursor.execute(Ax)   
   
    result=cursor.fetchone()  
    #print(result) 
    

    return result[0]

def createExcel(kurum,tesisAdi):
    workbook=Workbook()
    workbook.save(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi}.xlsx")
    print("Successfully created excel file")


def WriteToExcel(iladi,kurum,tesisAdi):
    connection=mysql.connector.connect(host="localhost", user= "root",password="password", database=f"{iladi}")

    workbook=openpyxl.load_workbook(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi}.xlsx") 
    sheets=workbook.sheetnames
    if "Sayfa1" in sheets:
        workbook["Sayfa1"].title=tesisAdi
    elif "Sheet" in sheets:
        workbook["Sheet"].title=tesisAdi
    else:
        workbook.create_sheet(title=f"{tesisAdi}")

   
    
    df=sql.read_sql("SELECT * FROM {tesisadi}".format(tesisadi=tesisAdi), connection)

    print(df)
    df.to_excel(f"C:/Users/{bilgisayarUsername}/Desktop/İÇME SUYU/{kurum}/{tesisAdi}.xlsx", sheet_name=tesisAdi)
    print("veri excele atıldı")
    
    time.sleep(1)

def valueColoring(iladi,kurum,tesisAdi):
    connection=mysql.connector.connect(host="localhost", user= "root",password="password", database=f"{iladi}")
    df=sql.read_sql("SELECT * FROM {tesisadi}".format(tesisadi=tesisAdi), connection)





  
            





    






